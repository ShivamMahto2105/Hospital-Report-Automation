import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

#Load The Dataset 
data = pd.read_csv("hospital_data.csv")

# Convert date columns
data["admission_date"] = pd.to_datetime(data["admission_date"])
data["discharge_date"] = pd.to_datetime(data["discharge_date"])

# Calculate hospital stay
data["stay_days"] = data["discharge_date"] - data["admission_date"]
data["stay_days"] = data["stay_days"].dt.days


#Basic Analysis

# Revenue by department
dept_revenue = data.groupby("department")["billing_amount"].sum()

# Insurance claim status count
claim_status = data["status"].value_counts()

# Monthly revenue
data["month"] = data["admission_date"].dt.month
monthly_revenue = data.groupby("month")["billing_amount"].sum()

# Top 10 patients with highest billing
top_patients = data.sort_values("billing_amount", ascending=False).head(10)


#Chart Creating 

# Chart 1 - Revenue by Department
dept_revenue.plot(kind="bar", title="Revenue by Department")
plt.ylabel("Revenue")
plt.savefig("dept_revenue.png")
plt.close()

# Chart 2 - Claim Status
claim_status.plot(kind="pie", autopct="%1.1f%%", title="Insurance Claim Status")
plt.savefig("claim_status.png")
plt.close()

# Chart 3 - Monthly Revenue
monthly_revenue.plot(kind="line", marker="o", title="Monthly Revenue")
plt.ylabel("Revenue")
plt.savefig("monthly_revenue.png")
plt.close()


#Excel Report 

with pd.ExcelWriter("hospital_report.xlsx") as writer:

    data.to_excel(writer, sheet_name="Raw Data", index=False)
    dept_revenue.to_excel(writer, sheet_name="Department Revenue")
    claim_status.to_excel(writer, sheet_name="Claim Status")
    monthly_revenue.to_excel(writer, sheet_name="Monthly Revenue")
    top_patients.to_excel(writer, sheet_name="Top Patients")

#Inserting Chart into Excel

workbook = load_workbook("hospital_report.xlsx")
chart_sheet = workbook.create_sheet("Charts")

chart1 = Image("dept_revenue.png")
chart2 = Image("claim_status.png")
chart3 = Image("monthly_revenue.png")

chart_sheet.add_image(chart1, "A1")
chart_sheet.add_image(chart2, "A20")
chart_sheet.add_image(chart3, "A40")

workbook.save("hospital_report.xlsx")


print("Report created successfully!")